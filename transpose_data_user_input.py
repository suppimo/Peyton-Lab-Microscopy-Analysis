import os
import itertools
import pandas as pd
import numpy as np 
import seaborn as sns 
import matplotlib.pyplot as plt
import xlrd 
import xlsxwriter
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException
import scipy
from scipy import stats

#TODO: Use os.path() often to make file reading and writing same as on mac and windows.

#------------------------------------------------------------------
#------------------------HELPERFUNCTIONS---------------------------
#------------------------------------------------------------------
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected


#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


#Create user_dict
def get_user_dict(conditions, timepoints, fluoros, xls, c_tp_list = None):
    #rename xls sheets to the user input? but you have to make sure they're the appropriate sheet

    #Get number of timepoints
    n_tps = len(timepoints)

    user_dict = {}

    #For each condition c,
    #for c in conditions:
        #Add the corresponding c_tp to the dictionary for that condition
        #Must change condition every n_tps entries
    for i in range(len(c_tp_list)):
        if i % n_tps == 0:
            temp_dic = {}
            c = conditions[i]
            #List of condition timepoints to be added to the condition dictionary entry
            copyList = c_tp_list[i : (i + (n_tps))] 
            for copy in copyList:
                #user_dict = {"TCPS" : {"TCPS Day 1" : dataframe tcps_d1}, "RGD" : {"RGD Day 1" : df rgd_d1}}
                temp_dic[copy] = pd.read_excel(xls, sheet_name = copy)#TODO: Should this be c_tp?
                user_dict[c] = temp_dic


    return user_dict

def parse_sheetnames(xls):
    #List of conditions from sheetnames
    c_list = []

    #List of timepoints from sheetnames
    tp_list = []

    #List of conditions and timepoints from sheetnames
    c_tp_list = []

    #Get sheetnames from excel
    sheetlist = xls.sheet_names

    #Parse sheetnames
    #For each sheetname,
    for sheet in sheetlist:
        #Remove leading and trailing spaces
        sheet = sheet.strip()
        #Add the c_tp to c_tp_list
        c_tp_list.append(sheet)
        #Parse the condition. Assume the condition and timepoint are separated by a space
        c = sheet.split(" ", 1)[0]
        c_list.append(c)
        #Parse the timepoint. Assume timepoint is an int
        sheetsplit = sheet.split()
        for i in sheetsplit:
            if not(str.isalpha(i)):
                tp = i
                tp_list.append(tp)

    return c_list, tp_list, c_tp_list

def get_longest_df(df_list):
    length = 0
    for i in range(len(df_list)):
        if df_list[i].shape[0] > length:
            index = i
    return index

#------------------------------------------------------------------------
#-----------------------------END HELPER FUNCTIONS-----------------------
#------------------------------------------------------------------------

#Step 1: Read TCPS, Matrigel, RGD, Lung, BM - Day 1, 7, 15 data
#This is the step where Django would be needed for Hyuna to upload
def user_input():
    """
    Prints program instructions and requirements.
    Requests user input for the following parameters:
    Name,
    Date of experiment,
    Experiment Conditions,
    Experiment Timepoints
    Fluorophores Used
    """

    #Get list of current files that will be analyzed:
    #get list of all files in cwd
    files = os.listdir(os.getcwd())
    #choose only the csv files
    files_csv = [f for f in files if f[-3:] == 'csv']


    #TODO: Instruct user where to put raw data files
    print("This tool takes raw data from an ImageJ mean fluorescence intensity analysis, ")
    print("transposes it, and generates correlative graphs from the given data. \n") 
    print("There are TWO important requirements to check for this program to function properly: \n")
    print("1: It is important that your files are .csv files, and are uploaded to the current working directory.\n")
    print("Your current working directory is: " + os.getcwd())
    print("Before you use this tool, make sure the raw data files you wish to analyze are uploaded to this location.\n")
    print("The files to be analyzed are: ")
    print(files_csv , "\n")
    print("2: Each sheet of your raw data must be labelled by timepoint and condition.")
    print("Otherwise, it is not possible to tell which file belongs to which condition and timepoint. \n")
    print("Example sheetnames: 'Matrigel Day 1', 'MG D1', 'RGD D4', 'Lung Day 7', etc. \n")
    print("This is to ensure the program can correctly identify each file's condition & timepoint. \n")
    print("Finally, this program automatically does not plot statistical outliers.\n")
    print("A statistical outlier has been defined as, for intensity data for one fluorophore, a data point\n")
    print("with an intensity greater than 2 standard deviations from the mean. Thus, data points with values\n")
    print("greater than 2 standard deviations from the mean intensity will not be plotted.\n")
    print("However, these data points are only excluded for graphing. They remain in the excel output of this program.")


    #Name
    name = input("Please enter your name: ")
    
    #Date of experiment
    date_of_experiment = input("\nPlease enter the date of experiment. Do not use slashes: ")
    #Catch typos
    if "/" in date_of_experiment:
        date_of_experiment = input("Your input contained a slash. Please reenter: ")
    
    #Experiment Timepoints
    print("\nPlease enter the timepoints of this experiment, each separated by a comma.")
    timepoints = input("Example: Day 1, Day 4, Day 7: ")
    #Convert timepoints to list: timepoints = ['Day 1', 'Day 4', 'Day 7']
    #Catch typos
    if not("," in timepoints):
        timepoints = input("Your input did not contain commas. Please reenter: ")
    timepoints = timepoints.split(",")
    #Remove leading and trailing spaces
    for i in range(0, len(timepoints)):
        timepoints[i] = timepoints[i].strip()
    #Store number of timepoints
    n_timepoints = len(timepoints)

    #Experimental conditions
    print("\nPlease enter the conditions of this experiment, each separated by a comma.")
    conditions = input("Example: TCPS, RGD, Lung, BM: ")
    #Convert conditions to list: timepoints = ['TCPS', 'RGD', 'Lung', 'BM']
    #Catch typos
    if not("," in conditions):
        conditions = input("Your input did not contain commas. Please reenter: ")
    conditions = conditions.split(",")
    #Remove leading and trailing spaces
    for i in range(len(conditions)):
        conditions[i] = conditions[i].strip()
    #Store number of conditions
    n_conditions = len(conditions)

    #Fluorophores
    print("\nPlease enter the fluorophores imaged, in the order that they were measured, each separated by a comma.")
    fluorophores = input("Example: Ki-67, AXL, E-Cadherin: ")
    #Convert fluorophores to list: fluorophores = ['Ki-67(Red)', 'AXL(Yellow)', E-Cadherin(Green)']
    #Catch typos
    if not("," in fluorophores):
        fluorophores = input("Your input did not contain commas. Please reenter: ")
    fluorophores = fluorophores.split(",")
    #Remove leading and trailing whitespaces
    for i in range(len(fluorophores)):
        fluorophores[i] = fluorophores[i].strip()
    #Store number of fluorophores
    n_fluorophores = len(fluorophores)

    return name, date_of_experiment, timepoints, n_timepoints, conditions, n_conditions, fluorophores, n_fluorophores

def load_files(name, date_of_experiment):
    #TODO: Make sure this works for ALL operating systems!
    file_location = os.getcwd() 
    file_name = f'{file_location}_{name}_{date_of_experiment}_compiled.xlsx'

    #Instantiate the writer for creating the transposed intermediate file
    writer = pd.ExcelWriter(file_name, engine = 'xlsxwriter')

    #get list of all files in the folder of choice
    files = os.listdir(file_location)
    #choose only the csv files
    files_csv = [f for f in files if f[-3:] == 'csv']

    #loop through all file names. Create a new excel where each sheet is the file data and the sheet name is file name
    for f in files_csv:
        df = pd.DataFrame(pd.read_csv(f))
        df = df['Mean']
        #remove '.csv' from each sheet name
        sheetname = f'{f}'
        sheetname = sheetname[:-4]
        #add each df to a new sheet in excel
        xls = df.to_excel(writer, sheet_name= sheetname)

    #Save the new excel file
    writer.save()
    return file_name, file_location, name, date_of_experiment

#Step 2: Compile all data frames into one master excel
def transpose_data(file_name, file_location, name, date_of_experiment, conditions, timepoints, fluorophores):
    #Load the excel file into memory
    #convert xlsx to xls so it can work with pandas
    workbook = openpyxl.load_workbook(file_name)
    outfile = f"{file_name.split('.')[0]}.xls"
    workbook.save(outfile)

    #open xls with pandas
    save_name = file_location + f"\{name}_{date_of_experiment}_transposed.xlsx"
    xls = pd.ExcelFile(file_name)

    #Parse sheetnames
    c_list, tp_list, c_tp_list = parse_sheetnames(xls)

    #Create dictionary of all permutations of timepoints and conditions:
    #dic = get_user_dict(conditions, timepoints, fluorophores, xls)
    dic = get_user_dict(c_list, timepoints, fluorophores, xls, c_tp_list)

    #n = how many rows to transpose
    n = len(fluorophores)

    #Reinstantiate writer for proper saving
    writer = pd.ExcelWriter(save_name, engine = 'xlsxwriter')

    #dic = {"TCPS" : {"TCPS Day 1" : dataframe tcps_d1}, "RGD" : {"RGD Day 1" : df rgd_d1}}
    for condition in dic: 
        #Get the value that is a dictionary, dic["TCPS"] = {"TCPS Day 1" : df tcps_d1, "TCPS Day 4" : df tcps_d4, ...}
        dic2 = dic[condition]
            #For each condition/timepoint pair,
        for cd_tp in dic2:
            #Retrieve the df
            df = dic2[cd_tp]
            #Transpose df and save
            if df.size % n == 0:
                #Transpose every nth row
                a = df['Mean'].to_numpy().reshape(-1, n)
                #Save transposed data as DF with column headers
                #TODO: Make headers a user input
                transposed_df = pd.DataFrame(a, columns = fluorophores)
            else:
                print(f'Number of entries is not divisible by number of rows in {cd_tp}. Please delete rows such that it is.')
    
            #Update transposed intermediate
            xls = transposed_df.to_excel(writer, sheet_name= f'{cd_tp}')
    
    #Save the new workbook
    writer.save()

    #Print sheetnames
    temp = pd.ExcelFile(f'{save_name}')
    print("Sheet names: " , temp.sheet_names)

    return save_name, n, dic, c_tp_list

def dump_to_template(fname, n):
    """
    Dumps the transposed intermediate
    to a template for further personal analyses.
    All three timepoints of each condition are on the same sheet now. 
    Further update: make the column headers the 
    different fluorophore names, and include the date and time. 
    """
    #Read Transposed Intermediate excel
    transposed_data = load_workbook(fname)

    #Create new workbook to reformat data into
    wb = Workbook()
    #Delete auto-generated Sheet 1
    wb.remove(wb['Sheet'])

    #For every n sheets:
    for sheet in transposed_data.worksheets:
        #Get sheet index: index starts at 0
        i = transposed_data.index(sheet)
        #print("Sheet index: ", i)

        #If you are on the first condition, 
        if (i % n == 0):         
            #Create sheetname for new workbook, remove last 3 characters to get rid of "D1" etc
            new_sheetname = f'{sheet}'[:-8]
            new_sheetname = new_sheetname[12:]
            #Create new sheet in new workbook
            wb.create_sheet(title = new_sheetname) 
            #Go from sheet i, to sheet i + (n-1)
            copyList = transposed_data.sheetnames[i : (i + (n))] 
        
            #Paste Start Column
            pasteCol = 1
            #For each sheet in transposed data while i % n == 1:
            for sheets in copyList:
                curSheet = transposed_data[sheets]
                #Get max column with data
                col = curSheet.max_column  
                #Get max row with data
                row = curSheet.max_row
                #Copy the data from [B1 to max col, max row]. Copy it from the appropriate sheet on transposed data
                copyData = copyRange(2, 1, col, row, curSheet)

                #Paste the copied data starting at column n +
                #print("paste start col = ", pasteCol) 
                #print("Sheet receiving: ", wb[new_sheetname])
                pasteRange(pasteCol, 1, n + pasteCol - 1, row, wb[new_sheetname], copyData)
                
                #Update paste col
                pasteCol += n

    #TODO: Remove this fname change. Just save as transposed when I have final version ready.
    fname = fname[:-5]
    fname = fname + "_pasted.xlsx"
    wb.save(fname)
    
    return wb

def generate_graphs(fname, fluoros, tps, conditions):
    """
    Takes transposed data and generates all sets of 
    dot plots and scatter plots across all conditions,
    fluorophores, and timepoints

    Loops through dic with conds and c_tp_list

    Extracts Data Frames from dic = get_user_dict(c_list, timepoints, fluorophores, xls, c_tp_list)

    df = dic[condition][cd_tp]
    """
    n = len(tps)

    #Retrieve data for each condition, timepoint, and fluorophore. 

    #Step 0: Generate scatter list of permutations of fluorophores to graph
    scatter_list = list(itertools.combinations(fluoros, 2))
    print("Fluorophores to be plotted against each other: ")
    for pair in scatter_list:
        print(pair[0], " vs. ", pair[1], "\n")
    
    #Step 1: Load the tranposed file
    #Read Transposed Intermediate excel
    #Returns a dictionary - the keys are the sheet names, and the values are the sheets as dataframes.

    #TODO: This has to be changed to the transposed intermediate
    #Need to move transposed intermediate to the cwd, or simply return it from the appropriate function
    df_dic = pd.read_excel(fname, sheet_name = None)

    xls = pd.ExcelFile(fname)

    n = len(tps)
    f = len(fluoros)
    c = len(conditions)

    #Create sheet of all dataframes that need a timepoint column
    #parse conditions from sheetnames
    sheet_conds, sheet_tps, sheet_cd_tps = parse_sheetnames(xls)
    sheet_conds = list(set(sheet_conds))

    #Step 2: Generate dictionary of all dfs grouped by condition
    #plot_dic = {'BM' : [bm1.df, bm2.df, ...], ...}
    plot_dic = {}
    for cond in sheet_conds:
        plot_dic[cond] = []

    keys = list(df_dic.keys())
    for cond in sheet_conds:
        for i in range(len(keys)):
            key_low = keys[i].lower()
            cond_low = cond.lower()
            if cond_low in key_low:
                plot_dic[cond].append(df_dic[keys[i]])
    
    #Step 3: Remove statistical outliers and add Timepoint column to each dataframe
    #For each condition in the dictionary,
    for cond in plot_dic:
        #Get the list of sheets for that condition
        sheets_list = plot_dic[cond]
        #for each df in a condition,
        for sheet_df in sheets_list:
            #Get column names
            columns = list(sheet_df.columns)
            #For each column name
            for col in columns:
                #Filter outliers by stddev in each column
                mean = sheet_df[col].mean()
                sd = sheet_df[col].std()
                sheet_df = sheet_df[(np.abs(sheet_df[col] - mean) < 2*sd )]
        #TODO: Add timepoint and drop "Unnamed: 0" column for each df in plot_dic
        for i in range(n):
            df = sheets_list[i]
            df['Timepoint'] = tps[i]
            #Remove unnamed column
            df.drop('Unnamed: 0', inplace=True, axis=1)

    #Step 4: Plot Scatter plots with or without trendline
    #TREND LINE OR NOT:
    finished = False
    while not(finished):
        answer = input("Would you like a trend line? Answer Y or N: ")
        if answer.upper() == "Y":
            #lmplot == scatter plot with trendline
            for cond in sheet_conds:
                for pair in scatter_list:    
                    g = sns.lmplot(x = pair[1], y = pair[0], hue = 'Timepoint', data = pd.concat(plot_dic[cond]), 
                    ci = None)#, col = f'{cond}')
                    plt.title(cond)
                    plt.show()
            finished = True
            break
        if answer.upper() == "N":
            #relplot == scatter plot without trendline
            for cond in sheet_conds:
                for pair in scatter_list:    
                    g = sns.relplot(x = pair[1], y = pair[0], hue = 'Timepoint', 
                    data = pd.concat(plot_dic[cond]), kind = 'scatter')#, col = f'{cond}')
                    plt.title(cond)
                    plt.show()
            finished = True
            break
        else:
            print("You did not type Y or N. Please reenter. \n")


    #Step 5: Restructure data frames for Dot Plots
    #Add condition column to each dataframe
    dotplot_df = pd.DataFrame()
    for cond in plot_dic:
        length = len(plot_dic[cond])
        for i in range(length):
            #Get dataframe
            cond_df = plot_dic[cond][i]
            #Add Condition column
            cond_df['Condition'] = cond
            #Concatenate the df to master dotplot df
            dotplot_df = pd.concat([dotplot_df, cond_df])

    #Step 6: Plot Dot Plots
    for f in fluoros:
        g = sns.stripplot(x = 'Condition', y = f, hue = "Timepoint", data = dotplot_df, jitter = True)
        plt.title(f)
        plt.show()


    pass


#Get user input
user_name, dOfE, tps, n_tps, conds, n_conds, fluoros, n_fluoros= user_input()
#Load all files
file_name, file_location, name, date_of_experiment = load_files(user_name, dOfE)
#Transpose Data
fname, n, dic, c_tp_list = transpose_data(file_name, file_location, name, date_of_experiment, conds, tps, fluoros)
#Format Transposed Data
dump_to_template(fname, n)
#Graph Transposed Data
#generate_graphs(dic, conds, c_tp_list)
generate_graphs(fname, fluoros, tps, conds)
print('All Done!')
